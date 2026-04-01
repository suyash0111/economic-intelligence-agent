"""
Excel Generator - Creates professionally formatted Excel spreadsheets.
Enhanced with importance scores, themes, auto-filter, and conditional formatting.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from pathlib import Path
from datetime import datetime
import logging

from collectors.base_collector import Article
from config.settings import Settings

logger = logging.getLogger(__name__)


class ExcelGenerator:
    """
    Generates the Master Intelligence Index - a professionally formatted
    Excel spreadsheet with all articles, scores, and AI analysis.
    """

    # Category colors
    CATEGORY_COLORS = {
        'Central Bank': 'FFEBEE',
        'International': 'E3F2FD',
        'Consulting': 'E8F5E9',
        'Think Tank': 'FFF3E0',
        'Investment Bank': 'F3E5F5',
        'News': 'FFFDE7',
        'Data Provider': 'E0F7FA',
        'Rating Agency': 'FCE4EC',
        'India': 'FFF8E1',
    }

    # Styles
    HEADER_FILL = PatternFill(start_color='0D1B2A', end_color='0D1B2A', fill_type='solid')
    HEADER_FONT = Font(color='FFFFFF', bold=True, size=10, name='Calibri')
    HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)

    BODY_FONT = Font(size=10, name='Calibri')
    BODY_ALIGN = Alignment(vertical='top', wrap_text=True)
    CENTER_ALIGN = Alignment(horizontal='center', vertical='center')

    THIN_BORDER = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )

    LINK_FONT = Font(color='0563C1', underline='single', size=9, name='Calibri')

    # Score colors for conditional formatting
    CRITICAL_FILL = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
    IMPORTANT_FILL = PatternFill(start_color='FFE0B2', end_color='FFE0B2', fill_type='solid')
    STANDARD_FILL = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')

    def __init__(self):
        self.wb = None
        self.ws = None

    def generate(
        self,
        articles: list[Article],
        date_range: str,
        output_path: Path = None
    ) -> Path:
        """Generate a professionally formatted Excel spreadsheet."""
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Master Intelligence Index"

        # Build the main sheet
        self._setup_headers()
        self._add_data_rows(articles)

        # Use actual row count (after filtering) for formatting
        actual_rows = self.ws.max_row - 1  # Subtract header row
        self._apply_formatting(actual_rows)
        self._add_conditional_formatting(actual_rows)
        self._add_auto_filter(actual_rows)

        # Summary sheet
        self._add_summary_sheet(articles, date_range)

        # Save
        if output_path is None:
            Settings.ensure_output_dir()
            output_path = Settings.OUTPUT_DIR / f"Master_Intelligence_Index_{datetime.now().strftime('%Y%m%d')}.xlsx"

        self.wb.save(output_path)
        logger.info(f"Excel saved: {output_path}")
        return output_path

    # =====================================================================
    # MAIN SHEET
    # =====================================================================

    def _setup_headers(self):
        """Set up column headers with enhanced columns."""
        headers = [
            ('Organization', 20),
            ('Title', 45),
            ('Published', 12),
            ('Org Category', 15),
            ('AI Category', 18),
            ('Type', 14),
            ('Importance', 11),
            ('Level', 10),
            ('Themes', 30),
            ('AI Summary', 70),
            ('AI Analysis', 80),
            ('Status', 12),
            ('Link', 50),
        ]

        for col, (header, width) in enumerate(headers, 1):
            cell = self.ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = self.HEADER_ALIGN
            cell.border = self.THIN_BORDER
            self.ws.column_dimensions[get_column_letter(col)].width = width

        # Freeze header row and first 2 columns
        self.ws.freeze_panes = 'C2'

    def _add_data_rows(self, articles: list[Article]):
        """Add data rows with all article fields, skipping filtered non-economic content."""
        # Filter out non-economic articles
        economic_articles = [a for a in articles if getattr(a, 'verification_status', '') != 'filtered']
        logger.info(f"Excel: {len(economic_articles)}/{len(articles)} articles after filtering non-economic content")

        for row_idx, article in enumerate(economic_articles, 2):
            # 1. Organization
            self.ws.cell(row=row_idx, column=1, value=article.source)

            # 2. Title
            self.ws.cell(row=row_idx, column=2, value=article.title)

            # 3. Published Date
            if article.published_date:
                cell = self.ws.cell(row=row_idx, column=3, value=article.published_date)
                cell.number_format = 'YYYY-MM-DD'
            else:
                self.ws.cell(row=row_idx, column=3, value='—')

            # 4. Organization Category
            self.ws.cell(row=row_idx, column=4, value=article.category)

            # 5. AI Category
            self.ws.cell(row=row_idx, column=5, value=article.ai_category or '—')

            # 6. Content Type
            self.ws.cell(row=row_idx, column=6,
                        value=article.content_type.replace('_', ' ').title())

            # 7. Importance Score
            self.ws.cell(row=row_idx, column=7, value=article.importance_score)

            # 8. Importance Level
            self.ws.cell(row=row_idx, column=8, value=article.importance_level or '—')

            # 9. Themes
            themes_text = ", ".join(article.themes) if article.themes else '—'
            self.ws.cell(row=row_idx, column=9, value=themes_text)

            # 10. AI Summary
            self.ws.cell(row=row_idx, column=10,
                        value=article.ai_summary or article.summary[:200] if article.summary else '—')

            # 11. AI Analysis (truncated for Excel)
            analysis_text = article.ai_analysis[:500] + "..." if article.ai_analysis and len(article.ai_analysis) > 500 else (article.ai_analysis or '—')
            self.ws.cell(row=row_idx, column=11, value=analysis_text)

            # 12. Verification Status
            status_labels = {
                'deep_verified': 'Deep Verified',
                'verified': 'Verified',
                'partial': 'Partial',
                'unverified': 'Unverified',
                'filtered': 'Filtered',
            }
            self.ws.cell(row=row_idx, column=12,
                        value=status_labels.get(article.verification_status, article.verification_status or '—'))

            # 13. Link (as hyperlink, with safety check for length)
            url = article.url
            link_cell = self.ws.cell(row=row_idx, column=13, value=url)
            if len(url) <= 255:
                try:
                    link_cell.hyperlink = url
                except Exception:
                    pass  # Skip hyperlink if URL is problematic
            link_cell.font = self.LINK_FONT

    def _apply_formatting(self, num_articles: int):
        """Apply row formatting with category-based coloring."""
        for row in range(2, num_articles + 2):
            org_name = self.ws.cell(row=row, column=1).value or ''
            color = self._get_org_color(org_name)
            fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

            for col in range(1, 14):
                cell = self.ws.cell(row=row, column=col)
                cell.border = self.THIN_BORDER
                cell.font = Font(size=10, name='Calibri')

                # Column-specific formatting
                if col in (10, 11):  # Summary and Analysis columns
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                elif col == 7:  # Importance Score
                    cell.alignment = self.CENTER_ALIGN
                    cell.font = Font(size=11, bold=True, name='Calibri')
                elif col == 8:  # Level
                    cell.alignment = self.CENTER_ALIGN
                    level = cell.value
                    if level == 'Critical':
                        cell.fill = self.CRITICAL_FILL
                        cell.font = Font(size=10, bold=True, color='B40000', name='Calibri')
                    elif level == 'Important':
                        cell.fill = self.IMPORTANT_FILL
                        cell.font = Font(size=10, bold=True, color='C86400', name='Calibri')
                    else:
                        cell.fill = self.STANDARD_FILL
                        cell.font = Font(size=10, color='007800', name='Calibri')
                    continue  # Skip category fill for this column
                elif col == 12:  # Status
                    cell.alignment = self.CENTER_ALIGN
                elif col == 13:  # Link
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                else:
                    cell.alignment = Alignment(vertical='center', wrap_text=True)

                # Apply category fill (except for Level column which has its own)
                if col != 8:
                    cell.fill = fill

        # Dynamic row heights based on content
        for row in range(2, num_articles + 2):
            summary_len = len(str(self.ws.cell(row=row, column=10).value or ''))
            if summary_len > 200:
                self.ws.row_dimensions[row].height = 80
            elif summary_len > 100:
                self.ws.row_dimensions[row].height = 55
            else:
                self.ws.row_dimensions[row].height = 40

    def _add_conditional_formatting(self, num_articles: int):
        """Add conditional formatting for importance scores."""
        score_range = f'G2:G{num_articles + 1}'

        # Critical: score >= 8 → red
        self.ws.conditional_formatting.add(
            score_range,
            CellIsRule(
                operator='greaterThanOrEqual',
                formula=['8'],
                fill=self.CRITICAL_FILL,
                font=Font(bold=True, color='B40000')
            )
        )

        # Important: score 5-7 → orange
        self.ws.conditional_formatting.add(
            score_range,
            CellIsRule(
                operator='between',
                formula=['5', '7'],
                fill=self.IMPORTANT_FILL,
                font=Font(bold=True, color='C86400')
            )
        )

        # Standard: score < 5 → green
        self.ws.conditional_formatting.add(
            score_range,
            CellIsRule(
                operator='lessThan',
                formula=['5'],
                fill=self.STANDARD_FILL,
                font=Font(color='007800')
            )
        )

    def _add_auto_filter(self, num_articles: int):
        """Add auto-filter to all columns."""
        last_col = get_column_letter(13)
        self.ws.auto_filter.ref = f'A1:{last_col}{num_articles + 1}'

    def _get_org_color(self, org_name: str) -> str:
        """Get the background color for an organization."""
        org_categories = {
            'Fed': 'Central Bank', 'ECB': 'Central Bank', 'BoE': 'Central Bank',
            'BoJ': 'Central Bank', 'PBoC': 'Central Bank',
            'IMF': 'International', 'World Bank': 'International', 'WTO': 'International',
            'OECD': 'International', 'WEF': 'International', 'BIS': 'International',
            'UNCTAD': 'International', 'ILO': 'International', 'UNDP': 'International',
            'NBER': 'International', 'ADB': 'International',
            'McKinsey': 'Consulting', 'Deloitte': 'Consulting', 'BCG': 'Consulting',
            'PwC': 'Consulting', 'EY': 'Consulting', 'KPMG': 'Consulting',
            'Accenture': 'Consulting', 'Bain': 'Consulting',
            'Brookings': 'Think Tank', 'PIIE': 'Think Tank', 'Bruegel': 'Think Tank',
            'Oxford Economics': 'Think Tank', 'Capital Economics': 'Think Tank',
            'Goldman Sachs': 'Investment Bank', 'JP Morgan': 'Investment Bank',
            'Bloomberg': 'News', 'Reuters': 'News', 'WSJ': 'News', 'FT': 'News',
            'S&P Global': 'Data Provider', 'EIU': 'Data Provider',
            "Moody's": 'Rating Agency', 'Fitch': 'Rating Agency', 'S&P Ratings': 'Rating Agency',
            'RBI': 'India', 'MoSPI': 'India', 'MoF India': 'India', 'NITI Aayog': 'India',
        }
        category = org_categories.get(org_name, '')
        return self.CATEGORY_COLORS.get(category, 'FFFFFF')

    # =====================================================================
    # SUMMARY SHEET
    # =====================================================================

    def _add_summary_sheet(self, articles: list[Article], date_range: str):
        """Add a formatted summary sheet with statistics."""
        ws = self.wb.create_sheet(title="Summary")

        # Title
        ws['A1'] = "Master Intelligence Index — Summary"
        ws['A1'].font = Font(bold=True, size=16, name='Calibri', color='0D1B2A')
        ws.merge_cells('A1:D1')

        ws['A2'] = f"Period: {date_range}"
        ws['A2'].font = Font(italic=True, size=12, name='Calibri', color='718096')
        ws.merge_cells('A2:D2')

        # ---- Overview Stats ----
        row = 4
        ws.cell(row=row, column=1, value="Overview").font = Font(bold=True, size=13, name='Calibri', color='0D1B2A')
        row += 1

        org_counts = {}
        for a in articles:
            org_counts[a.source] = org_counts.get(a.source, 0) + 1

        stats = [
            ("Total Articles/Reports", len(articles)),
            ("Organizations Covered", len(org_counts)),
            ("Critical Items (Score 8+)", sum(1 for a in articles if a.importance_score >= 8)),
            ("Important Items (Score 5-7)", sum(1 for a in articles if 5 <= a.importance_score <= 7)),
            ("Deep-Verified Reports", sum(1 for a in articles if a.verification_status == 'deep_verified')),
        ]

        for label, value in stats:
            ws.cell(row=row, column=1, value=label).font = Font(size=11, name='Calibri')
            val_cell = ws.cell(row=row, column=2, value=value)
            val_cell.font = Font(size=11, bold=True, name='Calibri')
            val_cell.alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=1).border = self.THIN_BORDER
            ws.cell(row=row, column=2).border = self.THIN_BORDER
            row += 1

        # ---- By AI Category ----
        row += 1
        ws.cell(row=row, column=1, value="By AI Category").font = Font(bold=True, size=13, name='Calibri', color='0D1B2A')
        row += 1

        # Header
        for i, h in enumerate(['Category', 'Count'], 1):
            cell = ws.cell(row=row, column=i, value=h)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = self.HEADER_ALIGN
            cell.border = self.THIN_BORDER
        row += 1

        cat_counts = {}
        for a in articles:
            cat = a.ai_category or a.category
            cat_counts[cat] = cat_counts.get(cat, 0) + 1

        for cat, count in sorted(cat_counts.items(), key=lambda x: -x[1]):
            ws.cell(row=row, column=1, value=cat).border = self.THIN_BORDER
            ws.cell(row=row, column=2, value=count).border = self.THIN_BORDER
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
            row += 1

        # ---- By Organization ----
        row += 1
        ws.cell(row=row, column=1, value="By Organization").font = Font(bold=True, size=13, name='Calibri', color='0D1B2A')
        row += 1

        for i, h in enumerate(['Organization', 'Count'], 1):
            cell = ws.cell(row=row, column=i, value=h)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = self.HEADER_ALIGN
            cell.border = self.THIN_BORDER
        row += 1

        for org, count in sorted(org_counts.items(), key=lambda x: -x[1]):
            ws.cell(row=row, column=1, value=org).border = self.THIN_BORDER
            ws.cell(row=row, column=2, value=count).border = self.THIN_BORDER
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 15
